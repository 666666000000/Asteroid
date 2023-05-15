#-*- coding : utf-8 -*-
import re
import qrcode
from openpyxl import load_workbook
from ..core import core

commands = {"qr"}
describe = "从文本、表格批量生成二维码"

def resolve(line,isReturn):
	arg,argLen = core.getArgList(line)
	return encode(arg,argLen)

def encode(arg,argLen):
	if argLen >= 3 and arg[1] == "f":
		src = core.getFilePathFromClipboard()
		if not src:
			print(f"源路径无效:{arg[1]}")
			return
		outDir = core.getOutputDirPath(arg[2],src[0],"file","first","checkDir","file")
		if not outDir:
			return
		encodeToFile(src,outDir,arg,argLen)
	else:
		return "continue"

def encodeToFile(src,dst,arg,argLen):
	for s in src:
		if s.endswith(".txt"):
			readTXT(s,dst)
		elif s.endswith(".xlsx"):
			p = [1,1,1,-1]
			for i in range(min(len(arg[3:]),len(p))):
				p[i] = int(arg[3+i])
			readXLSX(s,dst,p[0],p[1],p[2],p[3])

def readTXT(path,dst):
	with open(path, 'r',encoding = 'utf-8', errors = 'ignore') as file:
		for line in file:
			encodeStr(line.strip(),"None",dst)
		file.close()

def readXLSX(path,dst,strColumn,nameColumn,startRow,endRow):
	wb = load_workbook(path)
	ws = wb.active
	mrow = ws.max_row
	mcolumn = ws.max_column
	if strColumn > mcolumn or nameColumn > mcolumn:
		print("列参数错误")
		return
	if mrow > endRow and endRow != -1:
		mrow = endRow
	for x in range(startRow,mrow+1):
		src = ws.cell(row = x, column = strColumn).value
		name = ws.cell(row = x, column = nameColumn).value
		encodeStr(src,name,dst)

def encodeStr(src,name,dst):
	if not src:
		print("字符串错误")
		return
	img = qrcode.make(src)
	dst = core.replaceFileName(dst)
	if dst.find("*") != -1:
		if name == "None":
			dst = dst.replace("*",checkName(src)[0:127])
		else:
			dst = dst.replace("*",checkName(name)[0:127])
	print(f"保存至 :{dst}")
	img.save(dst)

def checkName(name):
	rstr = r'[\\/:*?"<>|\r\n]+'
	new_name = re.sub(rstr, "", name)
	return new_name.strip()