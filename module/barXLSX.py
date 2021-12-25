#-*- coding : utf-8 -*-
import re
from PIL import Image
from io import BytesIO
import barcode
from barcode.writer import ImageWriter
from openpyxl import load_workbook
from . import core

keywords = {"bar"}
describe = "从文本、表格批量生成条码"

def resolve(line):
	arg,argLen = core.getArgList(line)
	return encode(arg,argLen)

def encode(arg,argLen):
	global dstPath
	if argLen < 2:
		print("缺少参数")
		return
	if arg[1] not in barcode.PROVIDED_BARCODES:
		print(f"条码类型错误:{arg[1]}")
		print(barcode.PROVIDED_BARCODES)
		return
	if argLen >= 4 and arg[2] == "f":
		src = core.getFilePathFromClipboard()
		if not src:
			print(f"源路径无效:{arg[2]}")
			return
		outDir = core.getOutputDirPath(arg[3],src[0],"first","checkDir","file")
		if not outDir:
			return
		encodeToFile(arg[1],src,outDir,arg,argLen)
	else:
		return core.msg("continue")

def barCodeImage(type,input):
	output = BytesIO()
	EAN = barcode.get_barcode_class(type)
	try:
		EAN(input,writer=ImageWriter()).write(output)
	except:
		print("内容错误")
		return
	return Image.open(output)

def encodeToFile(codeType,src,dst,arg,argLen):
	for s in src:
		if s.endswith(".txt"):
			readTXT(codeType,s,dst)
		elif s.endswith(".xlsx"):
			strColumn = 1
			nameColumn = 1
			startRow = 1
			endRow = -1
			if argLen >= 5:
				strColumn = int(arg[4])
			if argLen >= 6:
				nameColumn = int(arg[5])
			if argLen >= 7:
				startRow = int(arg[6])
			if argLen >= 8:
				endRow = int(arg[7])
			readXLSX(codeType,s,dst,strColumn,nameColumn,startRow,endRow)

def readTXT(codeType,path,dst):
	with open(path, 'r',encoding = 'utf-8', errors = 'ignore') as file:
		for line in file:
			encodeStr(codeType,line.strip(),"None",dst)
		file.close()

def readXLSX(codeType,path,dst,strColumn,nameColumn,startRow,endRow):
	wb = load_workbook(path)
	ws = wb.active
	mrow = ws.max_row
	mcolumn = ws.max_column
	if strColumn > mcolumn or nameColumn > mcolumn:
		print("列参数错误")
		return
	if mrow > endRow and endRow != -1:
		mrow = endRow
	for x in range(startRow,mrow+1):
		src = ws.cell(row = x, column = strColumn).value
		name = ws.cell(row = x, column = nameColumn).value
		encodeStr(codeType,str(src).strip(),str(name).strip(),dst)

def encodeStr(codeType,src,name,dst):
	if not src:
		print("字符串错误")
		return
	img = barCodeImage(codeType,src)
	if not img:
		return
	dst = core.replaceFileName(dst)
	if dst.find("*") != -1:
		if name == "None":
			dst = dst.replace("*",checkName(src)[0:127])
		else:
			dst = dst.replace("*",checkName(name)[0:127])
	print(f"保存至 :{dst}")
	img.save(dst)

def checkName(name):
	rstr = r'[\\/:*?"<>|\r\n]+'
	new_name = re.sub(rstr, "", name)
	return new_name.strip()









